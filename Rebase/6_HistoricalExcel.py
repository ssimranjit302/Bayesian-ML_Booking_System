import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime, timezone, timedelta
from pathlib import Path
import glob
import os
from zoneinfo import ZoneInfo

# ---------- Step 1: Get London time ----------
london_time = datetime.now(ZoneInfo("Europe/London"))
london_time_str = london_time.strftime("%H-%M")
london_date_str = london_time.strftime("%Y-%m-%d")

# ---------- File paths ----------
BASE_DIR = Path(__file__).resolve().parent
input_file = BASE_DIR / "Master.json"
output_file = BASE_DIR / f"Rebase_Historical_{london_time.strftime('%Y-%m-%d')}_({london_time_str}).xlsx"

# ---------- Step 1.5: Remove previous Excel files ----------
today_pattern = str(BASE_DIR / f"Rebase_Historical_{london_date_str}_(*).xlsx")

for file_path in glob.glob(today_pattern):
    try:
        os.remove(file_path)
        print(f"Deleted previous file for today: {file_path}")
    except Exception as e:
        print(f"Error deleting {file_path}: {e}")

# ---------- Step 2: Load JSON data ----------
with open(input_file, "r") as f:
    data = json.load(f)

# ---------- Step 3: Convert to DataFrame ----------
df = pd.DataFrame(data)

# Select and order columns
columns = ["Date", "Start Time", "End Time", "Service/Class", "Staff/Suite.", "Capacity", "Booked"]
df = df[columns]

# ---------- Step 4: Save to Excel ----------
df.to_excel(output_file, index=False)

# ---------- Step 5: Merge Date cells ----------
wb = load_workbook(output_file)
ws = wb.active

merge_col = 1  # Column A = Date
start_row = 2  # Skip header
end_row = ws.max_row

current_date = ws.cell(row=start_row, column=merge_col).value
merge_start = start_row

for row in range(start_row + 1, end_row + 2):  # +2 to include last row
    cell_value = ws.cell(row=row, column=merge_col).value if row <= end_row else None
    if cell_value != current_date:
        if merge_start < row - 1:
            ws.merge_cells(start_row=merge_start, start_column=merge_col,
                           end_row=row - 1, end_column=merge_col)
            merged_cell = ws.cell(row=merge_start, column=merge_col)
            merged_cell.alignment = Alignment(vertical="center", horizontal="center")
        merge_start = row
        current_date = cell_value

# Adjust column widths
for col in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(col)].width = 18

# Save workbook
wb.save(output_file)

print(f" Excel file '{output_file}' created successfully with merged Date cells.")
