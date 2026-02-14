import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from pathlib import Path

# ---------- CONFIG ----------
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "RebelSlotsExcel.xlsx"

input_csv = BASE_DIR / "RebelSlots.csv"     # path to  CSV file
output_xlsx = BASE_DIR / "RebelSlotsExcel.xlsx" # output Excel file
# ----------------------------

def parse_class_info(class_str):
    if pd.isna(class_str) or " - " not in class_str:
        return None, None, None, None

    parts = [p.strip() for p in class_str.split("-")]
    if len(parts) < 4:
        return None, None, None, None

    Concept = parts[0].strip()
    start_time_str = parts[1].strip()
    duration_str = parts[2].replace("mins", "").replace("minutes", "").strip()
    trainer = parts[3].strip()

    # Convert start time to proper format
    try:
        start_time = datetime.strptime(start_time_str, "%H:%M").strftime("%I:%M %p")
    except:
        start_time = start_time_str  # fallback

    # Convert duration to integer minutes
    try:
        duration = int(duration_str)
    except:
        duration = None

    return Concept, start_time, duration, trainer


# Step 1: Read CSV
df = pd.read_csv(input_csv)

# Step 2: Extract info
records = []
for _, row in df.iterrows():
    date = row.get("date", "")
    club = row.get("club", "")
    booked = row.get("booked", "")
    capacity = row.get("capacity", "")
    class_str = row.get("new_page_time", "")

    Concept, start_time, duration, trainer = parse_class_info(class_str)

    # Compute End Time
    if start_time and duration:
        try:
            start_dt = datetime.strptime(start_time, "%I:%M %p")
            end_dt = start_dt + timedelta(minutes=duration)
            end_time = end_dt.strftime("%I:%M %p")
        except:
            end_time = ""
    else:
        end_time = ""

    records.append({
        "Date": date,
        "Start Time": start_time,
        "End Time": end_time,
        "Location/Club": club,
        "Concept Name": Concept,
        "Capacity": capacity,
        "Booked": booked,
        "Slots left": int(capacity) - int(booked)
    })

# Step 3: Create cleaned DataFrame
out_df = pd.DataFrame(records)

# Step 4: Save to Excel
out_df.to_excel(output_xlsx, index=False)

# Step 5: Merge cells for same Date
wb = load_workbook(output_xlsx)
ws = wb.active

# Find Date column index
date_col = 1
start_row = 2
end_row = ws.max_row

merge_start = start_row
current_date = ws.cell(row=start_row, column=date_col).value

for r in range(start_row + 1, end_row + 1):
    cell_date = ws.cell(row=r, column=date_col).value
    if cell_date != current_date:
        if merge_start != r - 1:
            ws.merge_cells(start_row=merge_start, start_column=date_col,
                           end_row=r - 1, end_column=date_col)
            ws.cell(row=merge_start, column=date_col).alignment = Alignment(vertical="center", horizontal="center")
        merge_start = r
        current_date = cell_date

# Merge last group
if merge_start != end_row:
    ws.merge_cells(start_row=merge_start, start_column=date_col,
                   end_row=end_row, end_column=date_col)
    ws.cell(row=merge_start, column=date_col).alignment = Alignment(vertical="center", horizontal="center")

wb.save(output_xlsx)

# Step 6: Save to JSON
json_output = BASE_DIR / "RebelSlotsData.json"
out_df.to_json(json_output, orient="records", indent=4, date_format="iso")
print(f" JSON data saved as: {json_output}")

print(f" Cleaned and formatted Excel saved as: {output_xlsx}")
