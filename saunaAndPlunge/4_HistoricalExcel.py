import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo
import glob
import os

# ---------- Get London time ----------
london_time = datetime.now(ZoneInfo("Europe/London"))
london_time_str = london_time.strftime("%H-%M")
london_date_str = london_time.strftime("%Y-%m-%d")

# ---------- Define file paths ----------
BASE_DIR = Path(__file__).resolve().parent
input_file = BASE_DIR / "Master.json"
output_file = BASE_DIR / f"Sauna&Plunge_Historical_{london_date_str}_({london_time_str}).xlsx"

# ----------Remove today's previous Excel files ----------
today_pattern = str(BASE_DIR / f"Sauna&Plunge_Historical_{london_date_str}_(*).xlsx")

for file_path in glob.glob(today_pattern):
    try:
        os.remove(file_path)
        print(f"Deleted previous file for today: {file_path}")
    except Exception as e:
        print(f"Error deleting {file_path}: {e}")

# ----------  Load JSON data ----------
with open(input_file, "r") as f:
    data = json.load(f)

# ----------  Convert to DataFrame ----------
df = pd.DataFrame(data, columns=[
    "Date", "Start Time", "End Time", "Service", "Booked", "Max Capacity"
])

# ---------- Write DataFrame to Excel ----------
df.to_excel(output_file, index=False, sheet_name="Sauna & Plunge")

# ---------- Merge same-date cells ----------
wb = load_workbook(output_file)
ws = wb.active

merge_col = 1  # Column A = Date
start_row = 2  # Skip header
end_row = ws.max_row

current_date = ws.cell(row=start_row, column=merge_col).value
merge_start = start_row

for row in range(start_row + 1, end_row + 2):  # +2 ensures last block is merged
    cell_value = ws.cell(row=row, column=merge_col).value if row <= end_row else None
    if cell_value != current_date:
        if merge_start < row - 1:
            ws.merge_cells(start_row=merge_start, start_column=merge_col,
                           end_row=row - 1, end_column=merge_col)
            merged_cell = ws.cell(row=merge_start, column=merge_col)
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        merge_start = row
        current_date = cell_value

# ---------- Auto-adjust column widths ----------
for col in range(1, ws.max_column + 1):
    max_length = 0
    column = get_column_letter(col)
    for cell in ws[column]:
        try:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 4  # add padding
    ws.column_dimensions[column].width = adjusted_width

# ----------  Save final workbook ----------
wb.save(output_file)

print(f"Excel file '{output_file.name}' created successfully with merged Date cells and auto-adjusted widths.")
